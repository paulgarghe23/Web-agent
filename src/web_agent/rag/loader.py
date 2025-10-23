from pathlib import Path
from typing import List, Tuple
import csv
import openpyxl
from pypdf import PdfReader
import logging
from web_agent.storage.r2_loader import download_files_from_r2

log = logging.getLogger(__name__)


def read_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def read_workstories_xlsx(path: Path) -> List[Tuple[str, str]]:
    """
    Lee únicamente la hoja 'Work Stories' y construye historias en formato:
    ("Story 1 - Snowball Collapse", "S: ... T: ... A: ... R: ...")
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Work Stories"] if "Work Stories" in wb.sheetnames else wb.active

    stories: List[Tuple[str, str]] = []
    current_title = ""
    buffer: list[str] = []

    def flush():
        nonlocal current_title, buffer
        if current_title and buffer:
            stories.append((current_title.strip(), " ".join(buffer).strip()))
        buffer = []

    for row in ws.iter_rows(values_only=True):
        # columna A = etiqueta (Story/S/T/A/R), columna B = texto principal
        a = (str(row[0]).strip() if row and row[0] is not None else "")
        b = (str(row[1]).strip() if row and len(row) > 1 and row[1] is not None else "")

        if not a and not b:
            continue  # fila vacía

        # Nuevo bloque de historia
        if a.lower().startswith("story"):
            flush()
            # título: "Story N - <B si existe>"
            current_title = f"{a}{' - ' + b if b else ''}"
            continue

        # Líneas STAR dentro de la historia (S/T/A/R en col A, texto en col B)
        if a in {"S", "T", "A", "R"} and b:
            buffer.append(f"{a}: {b}")
            continue

        # Otras filas informativas (por si hay subtítulos o notas en B)
        if b:
            buffer.append(b)

    flush()
    return stories


def load_docs(data_dir: str = "/tmp/data") -> List[Tuple[str, str]]:
    # First, download files from R2
    log.info("load_docs: downloading files from R2...")
    download_files_from_r2(target_dir=data_dir)
    
    log.info(f"load_docs: loading documents from {data_dir}")
    p = Path(data_dir)
    docs: List[Tuple[str, str]] = []

    # CV (ajusta el nombre si cambia)
    pdf = p / "Paul_G_CV.pdf"
    if pdf.exists():
        docs.append(("cv", read_pdf_text(pdf)))

    # Solo la hoja Work Stories
    xlsx = p / "work_stories.xlsx"
    if xlsx.exists():
        docs.extend(read_workstories_xlsx(xlsx))

    # Context.md
    context_md = p / "context.md"
    if context_md.exists():
        docs.append(("context", context_md.read_text(encoding="utf-8")))

    return docs